"""数据导入对话框 — 3步列选择向导"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional

import numpy as np
from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QAbstractItemView, QButtonGroup, QDialog, QFileDialog, QHBoxLayout, QHeaderView,
    QLabel, QSizePolicy, QStackedWidget,
    QTableWidgetItem, QVBoxLayout, QWidget,
)
from qfluentwidgets import (
    BodyLabel, CaptionLabel, ComboBox, InfoBar, InfoBarPosition,
    LineEdit, PrimaryPushButton, PushButton, RadioButton, SubtitleLabel, TableWidget,
)
from ui.widgets.focus_commit import install_click_away_focus_commit

from core.project_manager import project_manager
from models.schemas import DataSeries

_ROLES = ["X 轴", "Y 轴", "Y 误差棒", "X 误差棒", "跳过"]
SUPPORTED_IMPORT_SUFFIXES = (".csv", ".txt", ".dat", ".tsv", ".xlsx", ".xls", ".json", ".npy", ".npz")
_TEXT_PREVIEW_ENCODINGS = ("utf-8-sig", "utf-8", "gbk", "latin-1")


@dataclass
class PreviewParseResult:
    headers: List[str]
    rows: List[List[float]]
    total_rows: int
    detected_format: str
    column_types: List[str] = field(default_factory=list)
    encoding: str = ""
    delimiter: str = ""
    has_header: bool = False
    skipped_rows: int = 0
    applied_skip_rows: int = 0
    header_warning: str = ""
    sheet_names: List[str] = field(default_factory=list)
    selected_sheet: str = ""
    row_offset: int = 0
    row_limit: int = 80


@dataclass
class TextPreviewPage:
    content: str
    total_lines: int
    encoding: str
    line_offset: int
    line_limit: int


class ImportPreviewParser:
    @staticmethod
    def parse(file_path: str) -> tuple[List[str], List[List[float]]]:
        """解析文件，返回 (headers: List[str], rows: List[List[float]])。"""
        suffix = Path(file_path).suffix.lower()
        if suffix in (".xlsx", ".xls"):
            return _parse_excel(file_path)
        if suffix == ".json":
            return _parse_json(file_path)
        if suffix in (".npy", ".npz"):
            return _parse_npy(file_path)
        return _parse_csv(file_path)


def _parse_file_preview(file_path: str):
    return ImportPreviewParser.parse(file_path)


class ImportDialog(QDialog):
    """文件导入向导 — 三步流程：选文件 → 列角色 → 确认。"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("导入数据文件")
        self.setMinimumSize(760, 540)
        self.imported_series: List[DataSeries] = []
        self._import_completed = False

        self._file_path: str = ""
        self._preferred_file_name: str = ""
        self._raw_headers: List[str] = []
        self._raw_rows: List[List[float]] = []  # 全量数据行
        self._target_data_file_id: Optional[str] = None
        self._last_auto_data_file_name = ""
        self._click_away_focus_commit = install_click_away_focus_commit(self)

        self._setup_ui()

    def set_default_file_name(self, name: str) -> None:
        clean_name = str(name or "").strip()
        old_auto_name = self._default_data_file_name()
        self._preferred_file_name = self._normalize_default_file_name(clean_name)
        new_auto_name = self._default_data_file_name()
        if hasattr(self, "_data_file_name_edit"):
            current_name = self._data_file_name_edit.text().strip()
            if not current_name or current_name in {old_auto_name, self._last_auto_data_file_name}:
                self._data_file_name_edit.setText(new_auto_name)
        self._last_auto_data_file_name = new_auto_name

    def _setup_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(24, 20, 24, 16)
        root.setSpacing(12)

        # 步骤指示器
        self._step_label = BodyLabel("步骤 1 / 3：选择文件", self)
        root.addWidget(self._step_label)

        # 页面栈
        self._stack = QStackedWidget(self)
        root.addWidget(self._stack, 1)

        self._page1 = self._build_page1()
        self._page2 = self._build_page2()
        self._page3 = self._build_page3()
        self._stack.addWidget(self._page1)
        self._stack.addWidget(self._page2)
        self._stack.addWidget(self._page3)

        # 底部按钮
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        self._btn_back = PushButton("上一步", self)
        self._btn_back.setEnabled(False)
        self._btn_back.clicked.connect(self._go_back)
        btn_row.addWidget(self._btn_back)
        self._btn_next = PrimaryPushButton("下一步", self)
        self._btn_next.setEnabled(False)
        self._btn_next.clicked.connect(self._go_next)
        btn_row.addWidget(self._btn_next)
        self._btn_cancel = PushButton("取消", self)
        self._btn_cancel.clicked.connect(self.reject)
        btn_row.addWidget(self._btn_cancel)
        root.addLayout(btn_row)

    # ── Step 1 ──────────────────────────────────────────────────────────

    def _build_page1(self) -> QWidget:
        w = QWidget(self)
        lv = QVBoxLayout(w)
        lv.setSpacing(12)

        lv.addWidget(SubtitleLabel("选择数据文件", w))

        row = QHBoxLayout()
        self._path_edit = LineEdit(w)
        self._path_edit.setReadOnly(True)
        self._path_edit.setPlaceholderText("CSV / Excel / JSON / NumPy 文件…")
        self._path_edit.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        browse_btn = PushButton("浏览…", w)
        browse_btn.clicked.connect(self._browse)
        row.addWidget(self._path_edit)
        row.addWidget(browse_btn)
        lv.addLayout(row)

        self._p1_info = BodyLabel("", w)
        self._p1_info.setWordWrap(True)
        lv.addWidget(self._p1_info)
        lv.addStretch()
        return w

    def _browse(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择数据文件", "",
            "数据文件 (*.csv *.txt *.dat *.tsv *.xlsx *.xls *.json *.npy *.npz);;所有文件 (*)",
        )
        if not path:
            return
        self.load_file(path)

    def _clear_import_results(self) -> None:
        self.imported_series = []
        self._import_completed = False
        self._p3_summary.setText("")

    def _clear_column_assignment_state(self) -> None:
        self._col_table.clear()
        self._col_table.setRowCount(0)
        self._col_table.setColumnCount(0)
        self._name_edits = []
        self._role_buttons = []
        self._role_groups = []
        self._data_file_target_keys = [None]
        self._data_file_target_combo.blockSignals(True)
        self._data_file_target_combo.clear()
        self._data_file_target_combo.addItem("新建数据文件")
        self._data_file_target_combo.setCurrentIndex(0)
        self._data_file_target_combo.blockSignals(False)
        self._data_file_name_edit.clear()
        self._data_file_name_edit.setEnabled(True)
        self._data_file_target_hint.setText("可选择追加到现有数据文件，或新建一个数据文件承载本次导入结果。")
        self._target_data_file_id = None
        self._last_auto_data_file_name = ""

    def _show_file_selection_step(self, *, enable_next: bool) -> None:
        self._stack.setCurrentIndex(0)
        self._step_label.setText("步骤 1 / 3：选择文件")
        self._btn_back.setEnabled(False)
        self._btn_next.setText("下一步")
        self._btn_next.setEnabled(enable_next)

    def load_file(self, path: str) -> None:
        self._file_path = path
        self._path_edit.setText(path)
        self._raw_headers = []
        self._raw_rows = []
        self._p1_info.setText("")
        self._clear_import_results()
        self._clear_column_assignment_state()
        self._show_file_selection_step(enable_next=False)
        try:
            headers, rows = ImportPreviewParser.parse(path)
            self._raw_headers = headers
            self._raw_rows = rows
            n_rows = len(rows)
            n_cols = len(headers)
            self._p1_info.setText(f"✓ 检测到 {n_cols} 列，{n_rows} 行数据")
            self._btn_next.setEnabled(True)
        except Exception as e:
            self._p1_info.setText(f"⚠ 读取失败：{e}")
            self._btn_next.setEnabled(False)
            raise

    # ── Step 2 ──────────────────────────────────────────────────────────

    def _build_page2(self) -> QWidget:
        w = QWidget(self)
        lv = QVBoxLayout(w)
        lv.setSpacing(8)

        lv.addWidget(SubtitleLabel("分配列角色", w))
        lv.addWidget(BodyLabel("第一列可编辑变量名，第 2-6 列使用单选框指定导入角色。", w))

        self._col_table = TableWidget(w)
        self._col_table.setEditTriggers(TableWidget.EditTrigger.NoEditTriggers)
        self._col_table.verticalHeader().setVisible(False)
        self._col_table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self._col_table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self._col_table.setMouseTracking(False)
        self._col_table.viewport().setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self._col_table.viewport().setMouseTracking(False)
        self._col_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        for column in range(1, 1 + len(_ROLES)):
            self._col_table.horizontalHeader().setSectionResizeMode(column, QHeaderView.ResizeMode.ResizeToContents)
        lv.addWidget(self._col_table, 1)

        target_row = QHBoxLayout()
        target_row.addWidget(BodyLabel("目标数据文件:", w))
        self._data_file_target_combo = ComboBox(w)
        self._data_file_target_keys: List[Optional[str]] = []
        self._data_file_target_combo.currentIndexChanged.connect(self._on_data_file_target_changed)
        target_row.addWidget(self._data_file_target_combo, 1)
        self._data_file_name_edit = LineEdit(w)
        self._data_file_name_edit.setPlaceholderText("新数据文件名称，默认使用源文件名")
        target_row.addWidget(self._data_file_name_edit, 1)
        lv.addLayout(target_row)

        self._data_file_target_hint = CaptionLabel("可选择追加到现有数据文件，或新建一个数据文件承载本次导入结果。", w)
        self._data_file_target_hint.setWordWrap(True)
        lv.addWidget(self._data_file_target_hint)

        self._refresh_data_file_target_choices()
        return w

    def _populate_col_table(self):
        """用当前文件数据填充变量角色表格。"""
        if not self._raw_headers:
            return
        n_cols = len(self._raw_headers)
        self._col_table.clear()
        self._col_table.setColumnCount(1 + len(_ROLES))
        self._col_table.setHorizontalHeaderLabels(["变量名", *_ROLES])
        self._col_table.setRowCount(n_cols)

        self._name_edits: List[LineEdit] = []
        self._role_buttons: List[dict[str, RadioButton]] = []
        self._role_groups: List[QButtonGroup] = []

        for row in range(n_cols):
            self._col_table.setRowHeight(row, 42)

            name_edit = LineEdit(self._col_table)
            name_edit.setText(self._default_variable_name(row))
            name_edit.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
            self._col_table.setCellWidget(row, 0, name_edit)
            self._name_edits.append(name_edit)

            group = QButtonGroup(self._col_table)
            group.setExclusive(True)
            role_map: dict[str, RadioButton] = {}

            for column, role in enumerate(_ROLES, start=1):
                button = RadioButton("", self._col_table)
                button.setFocusPolicy(Qt.FocusPolicy.NoFocus)
                container = QWidget(self._col_table)
                container_layout = QHBoxLayout(container)
                container_layout.setContentsMargins(0, 0, 0, 0)
                container_layout.addStretch()
                container_layout.addWidget(button)
                container_layout.addStretch()
                self._col_table.setCellWidget(row, column, container)
                group.addButton(button)
                role_map[role] = button

            self._role_groups.append(group)
            self._role_buttons.append(role_map)

            default_role = "X 轴" if row == 0 else "Y 轴"
            role_map[default_role].setChecked(True)

        self._refresh_data_file_target_choices()

    def _default_variable_name(self, index: int) -> str:
        if index < len(self._raw_headers):
            header = str(self._raw_headers[index] or "").strip()
            if header and not re.fullmatch(r"col_\d+", header, flags=re.IGNORECASE):
                return header
        return f"变量{index + 1}"

    def _selected_roles(self) -> List[str]:
        roles: List[str] = []
        for role_map in getattr(self, "_role_buttons", []):
            role = next((name for name, button in role_map.items() if button.isChecked()), "跳过")
            roles.append(role)
        return roles

    def _variable_names(self) -> List[str]:
        names: List[str] = []
        for index, edit in enumerate(getattr(self, "_name_edits", [])):
            name = edit.text().strip()
            names.append(name or self._default_variable_name(index))
        return names

    def _existing_data_file_choices(self) -> List[tuple[str, str]]:
        project = project_manager.current_project
        if project is None:
            return []
        choices: List[tuple[str, str]] = []
        for data_file in getattr(project, "data_files", []):
            label = self._data_file_target_label(data_file.id)
            choices.append((label, data_file.id))
        return choices

    def _data_file_target_label(self, data_file_id: str) -> str:
        project = project_manager.current_project
        if project is None or project.tree is None:
            return f"数据文件 {data_file_id[:8]}"

        for node in getattr(project.tree, "nodes", []):
            if getattr(node, "kind", None) != "data_file":
                continue
            if getattr(node, "data_file_id", None) != data_file_id:
                continue
            label = project_manager.format_tree_path_label(node.id, separator="/", omit_root_group=True)
            if label:
                return label
            break

        data_file = project.find_data_file(data_file_id)
        if data_file is not None and data_file.name.strip():
            return data_file.name.strip()
        return f"数据文件 {data_file_id[:8]}"

    def _default_data_file_name(self) -> str:
        if self._preferred_file_name:
            return self._preferred_file_name
        if not self._file_path:
            return "导入数据"
        return self._normalize_default_file_name(Path(self._file_path).name) or "导入数据"

    @staticmethod
    def _normalize_default_file_name(name: str) -> str:
        clean_name = str(name or "").strip()
        if not clean_name:
            return ""
        suffix = Path(clean_name).suffix.lower()
        if suffix in SUPPORTED_IMPORT_SUFFIXES:
            stem = Path(clean_name).stem.strip()
            return stem or clean_name
        return clean_name

    def _refresh_data_file_target_choices(self) -> None:
        current_new_name = self._data_file_name_edit.text().strip() if hasattr(self, "_data_file_name_edit") else ""
        default_name = self._default_data_file_name()
        if not current_new_name or current_new_name == self._last_auto_data_file_name:
            current_new_name = default_name
        self._data_file_target_combo.blockSignals(True)
        self._data_file_target_combo.clear()
        self._data_file_target_keys = [None]
        self._data_file_target_combo.addItem("新建数据文件")
        for label, data_file_id in self._existing_data_file_choices():
            self._data_file_target_keys.append(data_file_id)
            self._data_file_target_combo.addItem(f"追加到现有: {label}")
        self._data_file_target_combo.setCurrentIndex(0)
        self._data_file_target_combo.blockSignals(False)
        self._last_auto_data_file_name = default_name
        self._data_file_name_edit.setText(current_new_name)
        self._on_data_file_target_changed(self._data_file_target_combo.currentIndex())

    def _current_target_data_file_id(self) -> Optional[str]:
        index = self._data_file_target_combo.currentIndex()
        if 0 <= index < len(self._data_file_target_keys):
            return self._data_file_target_keys[index]
        return None

    def _on_data_file_target_changed(self, _index: int) -> None:
        target_id = self._current_target_data_file_id()
        self._target_data_file_id = target_id
        create_new = target_id is None
        self._data_file_name_edit.setEnabled(create_new)
        if create_new:
            if not self._data_file_name_edit.text().strip():
                self._data_file_name_edit.setText(self._default_data_file_name())
            self._data_file_target_hint.setText("将导入结果写入一个新的数据文件，文件名可自定义。")
            return
        self._data_file_target_hint.setText("将把本次导入的系列追加到所选现有数据文件。")

    # ── Step 3 ──────────────────────────────────────────────────────────

    def _build_page3(self) -> QWidget:
        w = QWidget(self)
        lv = QVBoxLayout(w)
        lv.setSpacing(8)
        lv.addWidget(SubtitleLabel("确认导入", w))
        self._p3_summary = BodyLabel("", w)
        self._p3_summary.setWordWrap(True)
        lv.addWidget(self._p3_summary)
        lv.addStretch()
        return w

    # ── 导航 ─────────────────────────────────────────────────────────────

    def _go_next(self):
        cur = self._stack.currentIndex()
        if cur == 0:
            # Step1 → Step2: populate col table
            self._populate_col_table()
            self._stack.setCurrentIndex(1)
            self._step_label.setText("步骤 2 / 3：分配列角色")
            self._btn_back.setEnabled(True)
            self._btn_next.setText("导入")
        elif cur == 1:
            # Step2 → 执行导入 → Step3
            try:
                self.imported_series = self._do_import()
            except Exception as e:
                InfoBar.error(
                    title="导入失败", content=str(e),
                    position=InfoBarPosition.TOP, duration=4000, parent=self,
                )
                return
            summary = f"✓ 成功导入 {len(self.imported_series)} 条数据系列：\n"
            summary += "\n".join(f"  • {s.name} ({len(s.x)} 点)" for s in self.imported_series[:8])
            if len(self.imported_series) > 8:
                summary += f"\n  ... 共 {len(self.imported_series)} 条"
            self._p3_summary.setText(summary)
            self._stack.setCurrentIndex(2)
            self._step_label.setText("步骤 3 / 3：完成")
            self._btn_next.setText("完成")
            self._import_completed = True
        elif cur == 2 and self._import_completed:
            self.accept()

    def _go_back(self):
        cur = self._stack.currentIndex()
        if cur == 1:
            self._clear_import_results()
            self._show_file_selection_step(enable_next=bool(self._raw_headers and self._raw_rows))
        elif cur == 2:
            self._clear_import_results()
            self._stack.setCurrentIndex(1)
            self._step_label.setText("步骤 2 / 3：分配列角色")
            self._btn_back.setEnabled(True)
            self._btn_next.setText("导入")

    # ── 导入逻辑 ─────────────────────────────────────────────────────────

    def _do_import(self) -> List[DataSeries]:
        if not self._raw_rows or not self._raw_headers:
            raise ValueError("没有可用的数据")

        roles = self._selected_roles()
        variable_names = self._variable_names()

        x_cols = [i for i, r in enumerate(roles) if r == "X 轴"]
        y_cols = [i for i, r in enumerate(roles) if r == "Y 轴"]
        ye_cols = [i for i, r in enumerate(roles) if r == "Y 误差棒"]
        xe_cols = [i for i, r in enumerate(roles) if r == "X 误差棒"]

        if not x_cols:
            raise ValueError("至少需要指定一列作为 X 轴")
        if not y_cols:
            raise ValueError("至少需要指定一列作为 Y 轴")

        x_idx = x_cols[0]  # 取第一个 X 轴列
        arr = np.array(self._raw_rows, dtype=float)
        x_data = arr[:, x_idx].tolist()

        series_list: List[DataSeries] = []
        for y_idx in y_cols:
            y_data = arr[:, y_idx].tolist()
            col_name = variable_names[y_idx] if y_idx < len(variable_names) else self._default_variable_name(y_idx)
            s_name = col_name

            # 匹配 Y 误差棒（按照 y_cols 顺序对应 ye_cols）
            y_pos = y_cols.index(y_idx)
            y_err = None
            if ye_cols and y_pos < len(ye_cols):
                y_err = arr[:, ye_cols[y_pos]].tolist()

            series_list.append(DataSeries(
                name=s_name,
                x=x_data,
                y=y_data,
                y_err=y_err,
                x_label=variable_names[x_idx] if x_idx < len(variable_names) else "x",
                y_label=col_name,
                source="imported_file",
                source_file_path=str(self._file_path or ""),
            ))

        return series_list

    def import_with_default_options(self) -> List[DataSeries]:
        if not self._raw_headers or not self._raw_rows:
            if not self._file_path:
                raise ValueError("没有可用的数据")
            self.load_file(self._file_path)
        self._populate_col_table()
        self.imported_series = self._do_import()
        self._import_completed = True
        return list(self.imported_series)

    def get_results(self) -> List[DataSeries]:
        return self.imported_series

    def get_target_data_file_id(self) -> Optional[str]:
        return self._target_data_file_id

    def get_source_path(self) -> str:
        return self._file_path

    def get_file_name(self) -> str:
        return self._data_file_name_edit.text().strip() or self._default_data_file_name()


def analyze_file_preview(
    file_path: str,
    *,
    row_offset: int = 0,
    row_limit: int = 80,
    skip_rows: int = 0,
    sheet_name: str = "",
) -> PreviewParseResult:
    suffix = Path(file_path).suffix.lower()
    safe_offset = max(0, int(row_offset))
    safe_limit = max(1, int(row_limit))
    safe_skip_rows = max(0, int(skip_rows))
    if suffix in (".xlsx", ".xls"):
        return _analyze_excel_preview(
            file_path,
            row_offset=safe_offset,
            row_limit=safe_limit,
            skip_rows=safe_skip_rows,
            sheet_name=sheet_name,
        )
    if suffix == ".json":
        return _analyze_json_preview(file_path, row_offset=safe_offset, row_limit=safe_limit, skip_rows=safe_skip_rows)
    if suffix in (".npy", ".npz"):
        return _analyze_npy_preview(file_path, row_offset=safe_offset, row_limit=safe_limit, skip_rows=safe_skip_rows)
    return _analyze_csv_preview(file_path, row_offset=safe_offset, row_limit=safe_limit, skip_rows=safe_skip_rows)


def get_excel_sheet_names(file_path: str) -> List[str]:
    suffix = Path(file_path).suffix.lower()
    if suffix not in {".xlsx", ".xls"}:
        return []
    import openpyxl

    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        return [str(name) for name in workbook.sheetnames]
    finally:
        workbook.close()


def read_text_preview_page(
    file_path: str,
    *,
    line_offset: int = 0,
    line_limit: int = 80,
) -> TextPreviewPage:
    encoding = _detect_text_encoding(file_path)
    safe_offset = max(0, int(line_offset))
    safe_limit = max(1, int(line_limit))
    lines: List[str] = []
    total_lines = 0
    with open(file_path, encoding=encoding, errors="replace", newline="") as handle:
        for raw_line in handle:
            if safe_offset <= total_lines < safe_offset + safe_limit:
                lines.append(raw_line.rstrip("\r\n"))
            total_lines += 1
    return TextPreviewPage(
        content="\n".join(lines),
        total_lines=total_lines,
        encoding=encoding,
        line_offset=safe_offset,
        line_limit=safe_limit,
    )


def _detect_text_encoding(file_path: str) -> str:
    for encoding in _TEXT_PREVIEW_ENCODINGS:
        try:
            with open(file_path, encoding=encoding, newline="") as handle:
                for _ in range(32):
                    if handle.readline() == "":
                        break
            return encoding
        except UnicodeDecodeError:
            continue
    raise ValueError("无法识别文本编码，请尝试转换为 UTF-8 后重试")


def _iter_filtered_text_lines(file_path: str, encoding: str, *, skip_rows: int = 0):
    skipped = 0
    with open(file_path, encoding=encoding, newline="") as handle:
        for raw_line in handle:
            line = raw_line.rstrip("\r\n")
            if not line.strip():
                continue
            if line.lstrip().startswith(("#", "%", "!", "/")):
                continue
            if skipped < skip_rows:
                skipped += 1
                continue
            yield line


def _detect_delimiter(lines: List[str]):
    joined = "\n".join(lines[:5])
    if "\t" in joined:
        return "\t"
    comma_score = sum(line.count(",") for line in lines[:5])
    semicolon_score = sum(line.count(";") for line in lines[:5])
    if comma_score == 0 and semicolon_score == 0:
        return None
    return "," if comma_score >= semicolon_score else ";"


def _describe_delimiter(delimiter) -> str:
    if delimiter == "\t":
        return "Tab"
    if delimiter == ",":
        return "逗号 (,)"
    if delimiter == ";":
        return "分号 (;)"
    return "空白字符"


def _split_text_fields(line: str, delimiter):
    if delimiter is None:
        return re.split(r"\s+", line.strip())
    return line.split(delimiter)


def _looks_like_numeric_row(values: List[str]) -> bool:
    try:
        [float(value.strip()) for value in values if value.strip()]
        return True
    except ValueError:
        return False


def _build_column_types(rows: List[List[float]], ncols: int) -> List[str]:
    seen = [False] * ncols
    has_fraction = [False] * ncols
    for row in rows:
        if len(row) != ncols:
            continue
        for column_index, value in enumerate(row):
            seen[column_index] = True
            if not float(value).is_integer():
                has_fraction[column_index] = True
    result: List[str] = []
    for column_index in range(ncols):
        if not seen[column_index]:
            result.append("空列")
        elif has_fraction[column_index]:
            result.append("浮点数")
        else:
            result.append("整数")
    return result


def _update_column_type_flags(seen: List[bool], has_fraction: List[bool], row: List[float]) -> None:
    for column_index, value in enumerate(row):
        seen[column_index] = True
        if not float(value).is_integer():
            has_fraction[column_index] = True


def _column_types_from_flags(seen: List[bool], has_fraction: List[bool]) -> List[str]:
    result: List[str] = []
    for column_index in range(len(seen)):
        if not seen[column_index]:
            result.append("空列")
        elif has_fraction[column_index]:
            result.append("浮点数")
        else:
            result.append("整数")
    return result


def _iter_excel_non_empty_rows(worksheet, *, skip_rows: int = 0):
    skipped = 0
    for row in worksheet.iter_rows(values_only=True):
        values = list(row)
        if any(value not in (None, "") for value in values):
            if skipped < skip_rows:
                skipped += 1
                continue
            yield values


def _coerce_excel_numeric_row(row) -> Optional[List[float]]:
    values: List[float] = []
    for value in row:
        if value in (None, ""):
            continue
        try:
            values.append(float(value))
        except (TypeError, ValueError):
            return None
    return values or None


def _analyze_csv_preview(file_path: str, *, row_offset: int, row_limit: int, skip_rows: int) -> PreviewParseResult:
    encoding = _detect_text_encoding(file_path)
    sample_lines: List[str] = []
    for line in _iter_filtered_text_lines(file_path, encoding, skip_rows=skip_rows):
        sample_lines.append(line)
        if len(sample_lines) >= 5:
            break
    if not sample_lines:
        raise ValueError("文件为空，或可解析内容为空")

    delimiter = _detect_delimiter(sample_lines)
    first_fields = _split_text_fields(sample_lines[0], delimiter)
    has_header = False if skip_rows > 0 else not _looks_like_numeric_row(first_fields)
    start_index = 1 if has_header else 0

    length_counts: dict[int, int] = {}
    invalid_rows = 0
    for line_index, line in enumerate(_iter_filtered_text_lines(file_path, encoding, skip_rows=skip_rows)):
        if line_index < start_index:
            continue
        parts = _split_text_fields(line, delimiter)
        try:
            row = [float(value) for value in parts if value.strip()]
        except ValueError:
            invalid_rows += 1
            continue
        if not row:
            continue
        length_counts[len(row)] = length_counts.get(len(row), 0) + 1

    if not length_counts:
        raise ValueError("未找到可导入的数值行，请检查编码、分隔符或表头设置")

    ncols = max(length_counts, key=lambda column_count: (length_counts[column_count], column_count))
    headers = [value.strip().strip('"\'') for value in first_fields]
    header_warning = ""
    if has_header and (len(headers) != ncols or any(not header for header in headers)):
        header_warning = "表头缺项或错误"
    if not has_header or header_warning:
        headers = [f"col_{index}" for index in range(ncols)]

    rows: List[List[float]] = []
    seen_types = [False] * ncols
    has_fraction = [False] * ncols
    total_rows = 0
    for line_index, line in enumerate(_iter_filtered_text_lines(file_path, encoding, skip_rows=skip_rows)):
        if line_index < start_index:
            continue
        parts = _split_text_fields(line, delimiter)
        try:
            row = [float(value) for value in parts if value.strip()]
        except ValueError:
            continue
        if not row or len(row) != ncols:
            continue
        _update_column_type_flags(seen_types, has_fraction, row)
        if row_offset <= total_rows < row_offset + row_limit:
            rows.append(row)
        total_rows += 1

    return PreviewParseResult(
        headers=headers,
        rows=rows,
        total_rows=total_rows,
        detected_format="文本表格",
        column_types=_column_types_from_flags(seen_types, has_fraction),
        encoding=encoding,
        delimiter=_describe_delimiter(delimiter),
        has_header=has_header,
        skipped_rows=invalid_rows,
        applied_skip_rows=skip_rows,
        header_warning=header_warning,
        row_offset=row_offset,
        row_limit=row_limit,
    )


def _analyze_excel_preview(
    file_path: str,
    *,
    row_offset: int,
    row_limit: int,
    skip_rows: int,
    sheet_name: str,
) -> PreviewParseResult:
    import openpyxl

    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        sheet_names = [str(name) for name in workbook.sheetnames]
        if not sheet_names:
            raise ValueError("Excel 工作簿中没有可用工作表")
        selected_sheet = sheet_name if sheet_name in sheet_names else sheet_names[0]
        worksheet = workbook[selected_sheet]

        row_iterator = _iter_excel_non_empty_rows(worksheet, skip_rows=skip_rows)
        try:
            first_row = next(row_iterator)
        except StopIteration as exc:
            raise ValueError(f"工作表“{selected_sheet}”为空") from exc

        first_values = [str(value).strip() if value is not None else "" for value in first_row]
        has_header = False if skip_rows > 0 else not _looks_like_numeric_row(first_values)
        length_counts: dict[int, int] = {}
        invalid_rows = 0

        for row_index, row in enumerate(_iter_excel_non_empty_rows(worksheet, skip_rows=skip_rows)):
            if has_header and row_index == 0:
                continue
            numeric_row = _coerce_excel_numeric_row(row)
            if numeric_row is None:
                invalid_rows += 1
                continue
            length_counts[len(numeric_row)] = length_counts.get(len(numeric_row), 0) + 1

        if not length_counts:
            raise ValueError(f"工作表“{selected_sheet}”中未找到可导入的数值数据")

        ncols = max(length_counts, key=lambda column_count: (length_counts[column_count], column_count))
        raw_headers = [str(value).strip() if value not in (None, "") else "" for value in first_row]
        header_warning = ""
        if has_header and (len(raw_headers) != ncols or any(not header for header in raw_headers)):
            header_warning = "表头缺项或错误"
        headers = raw_headers
        if not has_header or header_warning:
            headers = [f"col_{index}" for index in range(ncols)]

        rows: List[List[float]] = []
        seen_types = [False] * ncols
        has_fraction = [False] * ncols
        total_rows = 0
        for row_index, row in enumerate(_iter_excel_non_empty_rows(worksheet, skip_rows=skip_rows)):
            if has_header and row_index == 0:
                continue
            numeric_row = _coerce_excel_numeric_row(row)
            if numeric_row is None or len(numeric_row) != ncols:
                continue
            _update_column_type_flags(seen_types, has_fraction, numeric_row)
            if row_offset <= total_rows < row_offset + row_limit:
                rows.append(numeric_row)
            total_rows += 1

        return PreviewParseResult(
            headers=headers,
            rows=rows,
            total_rows=total_rows,
            detected_format="Excel",
            column_types=_column_types_from_flags(seen_types, has_fraction),
            encoding="OpenXML",
            delimiter="工作表结构",
            has_header=has_header,
            skipped_rows=invalid_rows,
            applied_skip_rows=skip_rows,
            header_warning=header_warning,
            sheet_names=sheet_names,
            selected_sheet=selected_sheet,
            row_offset=row_offset,
            row_limit=row_limit,
        )
    finally:
        workbook.close()


def _analyze_json_preview(file_path: str, *, row_offset: int, row_limit: int, skip_rows: int) -> PreviewParseResult:
    headers, rows = _parse_json(file_path)
    safe_skip_rows = min(max(0, skip_rows), len(rows))
    available_rows = rows[safe_skip_rows:]
    ncols = len(headers)
    return PreviewParseResult(
        headers=headers,
        rows=available_rows[row_offset: row_offset + row_limit],
        total_rows=len(available_rows),
        detected_format="JSON",
        column_types=_build_column_types(rows, ncols),
        encoding="utf-8",
        delimiter="键结构",
        has_header=True,
        applied_skip_rows=safe_skip_rows,
        row_offset=row_offset,
        row_limit=row_limit,
    )


def _analyze_npy_preview(file_path: str, *, row_offset: int, row_limit: int, skip_rows: int) -> PreviewParseResult:
    headers, rows = _parse_npy(file_path)
    safe_skip_rows = min(max(0, skip_rows), len(rows))
    available_rows = rows[safe_skip_rows:]
    ncols = len(headers)
    return PreviewParseResult(
        headers=headers,
        rows=available_rows[row_offset: row_offset + row_limit],
        total_rows=len(available_rows),
        detected_format="NumPy",
        column_types=_build_column_types(rows, ncols),
        encoding="二进制数组",
        delimiter="数组列",
        has_header=True,
        applied_skip_rows=safe_skip_rows,
        row_offset=row_offset,
        row_limit=row_limit,
    )


def _parse_csv(file_path: str):
    raw_lines: List[str] = []
    for enc in _TEXT_PREVIEW_ENCODINGS:
        try:
            with open(file_path, encoding=enc, newline="") as f:
                raw_lines = f.readlines()
            break
        except UnicodeDecodeError:
            continue
    data_lines = [l.rstrip("\r\n") for l in raw_lines
                  if l.strip() and not l.lstrip().startswith(("#", "%", "!", "/"))]
    if not data_lines:
        raise ValueError("文件为空")

    sep = _detect_delimiter(data_lines[:5])

    def split(line):
        return _split_text_fields(line, sep)

    headers: Optional[List[str]] = None
    start = 0
    first = split(data_lines[0])
    try:
        [float(v.strip()) for v in first if v.strip()]
    except ValueError:
        headers = [v.strip().strip('"\'') for v in first]
        start = 1

    rows: List[List[float]] = []
    for line in data_lines[start:]:
        parts = split(line)
        try:
            row = [float(v) for v in parts if v.strip()]
            if row:
                rows.append(row)
        except ValueError:
            continue
    if not rows:
        raise ValueError("未找到数值数据")

    ncols = max(set(len(r) for r in rows), key=lambda c: sum(1 for r in rows if len(r) == c))
    rows = [r for r in rows if len(r) == ncols]
    if headers is None or len(headers) != ncols:
        headers = [f"col_{i}" for i in range(ncols)]
    return headers, rows


def _parse_excel(file_path: str):
    import openpyxl
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    raw = list(ws.iter_rows(values_only=True))
    wb.close()
    if not raw:
        raise ValueError("Excel 工作表为空")
    first = raw[0]
    try:
        [float(v) for v in first if v is not None]
        headers = [f"col_{i}" for i in range(len(first))]
        data_rows = raw
    except (ValueError, TypeError):
        headers = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(first)]
        data_rows = raw[1:]

    rows: List[List[float]] = []
    for row in data_rows:
        try:
            r = [float(v) for v in row if v is not None]
            if r:
                rows.append(r)
        except (ValueError, TypeError):
            continue
    if not rows:
        raise ValueError("Excel 中无数值数据")
    ncols = max(set(len(r) for r in rows), key=lambda c: sum(1 for r in rows if len(r) == c))
    rows = [r for r in rows if len(r) == ncols]
    if len(headers) != ncols:
        headers = [f"col_{i}" for i in range(ncols)]
    return headers, rows


def _parse_json(file_path: str):
    import json
    with open(file_path, encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, dict):
        # {x:[...], y:[...]} 格式
        keys = list(data.keys())
        rows = list(zip(*[data[k] for k in keys]))
        rows_f = [[float(v) for v in row] for row in rows]
        return keys, rows_f
    if isinstance(data, list) and data and isinstance(data[0], dict):
        keys = list(data[0].keys())
        rows_f = []
        for item in data:
            try:
                rows_f.append([float(item.get(k, float("nan"))) for k in keys])
            except Exception:
                continue
        return keys, rows_f
    raise ValueError("无法识别 JSON 结构")


def _parse_npy(file_path: str):
    arr = np.load(file_path, allow_pickle=False)
    if arr.ndim == 1:
        arr = arr.reshape(-1, 1)
    if arr.ndim != 2:
        raise ValueError("NumPy 数组应为 1D 或 2D")
    headers = [f"col_{i}" for i in range(arr.shape[1])]
    rows = arr.tolist()
    return headers, rows
