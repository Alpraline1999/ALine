"""
数据操作层 — 文件导入 / DataSeries 创建

职责：将外部文件（CSV/Excel/JSON/NumPy）解析为 DataSeries 列表。
UI 层不直接解析文件，通过此模块调用。
"""
from __future__ import annotations

import json
import os
from pathlib import Path
from typing import List, Optional

from models.schemas import Curve, DataSeries


# ──────────────────────────────────────────────────────────
# 单文件导入 → List[DataSeries]
# ──────────────────────────────────────────────────────────

def import_csv(file_path: str) -> List[DataSeries]:
    """从 CSV / TXT / DAT / TSV 导入，每对相邻列自动生成一条 DataSeries。

    - 自动检测分隔符
    - 首行含非数字时视为表头
    - 多列文件：以首列为 x，其余列各生成一条系列；或相邻两列配对
    Returns:
        List[DataSeries]，至少一个元素。
    Raises:
        FileNotFoundError / ValueError
    """
    file_path = str(Path(file_path).expanduser().resolve())
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")

    with open(file_path, "r", encoding="utf-8-sig", errors="replace") as f:
        lines = [l.rstrip("\n") for l in f]

    data_lines = [l for l in lines if l.strip() and not l.strip().startswith("#")]
    if not data_lines:
        raise ValueError("文件为空或仅含注释")

    sep = _detect_sep(data_lines[0])

    def split(line: str) -> List[str]:
        return line.split() if sep is None else [c.strip() for c in line.split(sep)]

    first = split(data_lines[0])
    has_header = _is_header(first)
    headers = first if has_header else [f"col_{i}" for i in range(len(first))]
    data_rows = data_lines[1:] if has_header else data_lines

    n = len(headers)
    cols: List[List[float]] = [[] for _ in range(n)]
    for line in data_rows:
        if not line.strip():
            continue
        cells = split(line)
        for i in range(n):
            try:
                cols[i].append(float(cells[i]) if i < len(cells) else float("nan"))
            except ValueError:
                cols[i].append(float("nan"))

    stem = Path(file_path).stem
    return _cols_to_series(cols, headers, stem)


def import_excel(file_path: str, sheet: Optional[str] = None) -> List[DataSeries]:
    """从 Excel 导入（需要 openpyxl）。"""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("请先安装 openpyxl：uv pip install openpyxl")

    fp = str(Path(file_path).expanduser().resolve())
    if not os.path.exists(fp):
        raise FileNotFoundError(f"文件不存在: {fp}")

    wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        raise ValueError("Excel 工作表为空")

    headers = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(rows[0])]
    n = len(headers)
    cols: List[List[float]] = [[] for _ in range(n)]
    for row in rows[1:]:
        for i in range(n):
            v = row[i] if i < len(row) else None
            try:
                cols[i].append(float(v) if v is not None else float("nan"))
            except (ValueError, TypeError):
                cols[i].append(float("nan"))

    stem = Path(fp).stem
    return _cols_to_series(cols, headers, stem)


def import_json(file_path: str) -> List[DataSeries]:
    """从 JSON 导入。支持 [{x,y},...] 或 {x:[...],y:[...]} 格式。"""
    fp = str(Path(file_path).expanduser().resolve())
    if not os.path.exists(fp):
        raise FileNotFoundError(f"文件不存在: {fp}")
    with open(fp, "r", encoding="utf-8") as f:
        data = json.load(f)
    stem = Path(fp).stem

    if isinstance(data, list):
        if not data:
            raise ValueError("JSON 数组为空")
        keys = list(data[0].keys())
        cols: dict[str, list[float]] = {k: [] for k in keys}
        for item in data:
            for k in keys:
                try:
                    cols[k].append(float(item.get(k, float("nan"))))
                except (ValueError, TypeError):
                    cols[k].append(float("nan"))
        col_lists = [cols[k] for k in keys]
        return _cols_to_series(col_lists, keys, stem)

    elif isinstance(data, dict):
        keys = list(data.keys())
        col_lists = []
        for k in keys:
            v = data[k]
            if isinstance(v, list):
                vals = []
                for item in v:
                    try:
                        vals.append(float(item))
                    except (ValueError, TypeError):
                        vals.append(float("nan"))
                col_lists.append(vals)
        return _cols_to_series(col_lists, keys, stem)

    raise ValueError("不支持的 JSON 结构")


def import_numpy(file_path: str) -> List[DataSeries]:
    """从 NumPy .npy / .npz 导入（需要 numpy）。"""
    try:
        import numpy as np
    except ImportError:
        raise ImportError("请先安装 numpy")

    fp = str(Path(file_path).expanduser().resolve())
    if not os.path.exists(fp):
        raise FileNotFoundError(f"文件不存在: {fp}")

    suffix = Path(fp).suffix.lower()
    stem = Path(fp).stem

    if suffix == ".npy":
        arr = np.load(fp, allow_pickle=False)
        if arr.ndim == 1:
            return _cols_to_series([arr.tolist()], ["col_0"], stem)
        elif arr.ndim == 2:
            cols = [arr[:, i].tolist() for i in range(arr.shape[1])]
            headers = [f"col_{i}" for i in range(arr.shape[1])]
            return _cols_to_series(cols, headers, stem)
        raise ValueError("只支持 1D 或 2D NumPy 数组")

    elif suffix == ".npz":
        data = np.load(fp)
        cols, headers = [], []
        for k in data.files:
            arr = data[k]
            if arr.ndim == 1:
                cols.append(arr.tolist())
                headers.append(k)
        return _cols_to_series(cols, headers, stem)

    raise ValueError(f"不支持的 NumPy 格式: {suffix}")


def import_file(file_path: str) -> List[DataSeries]:
    """按扩展名自动选择导入器。"""
    suffix = Path(file_path).suffix.lower()
    if suffix in (".xlsx", ".xls"):
        return import_excel(file_path)
    elif suffix == ".json":
        return import_json(file_path)
    elif suffix in (".npy", ".npz"):
        return import_numpy(file_path)
    else:
        return import_csv(file_path)


# ──────────────────────────────────────────────────────────
# PyLine Curve → DataSeries
# ──────────────────────────────────────────────────────────

def curve_to_series(curve: Curve, x_label: str = "x", y_label: str = "y") -> DataSeries:
    """将 PyLine Curve 的真实坐标复制为独立 DataSeries。"""
    calib = curve.calibration
    if calib and calib.coord_type == "polar":
        x_label = x_label if x_label != "x" else "r"
        y_label = y_label if y_label != "y" else "θ"
    return DataSeries(
        name=curve.name,
        x=list(curve.x_actual),
        y=list(curve.y_actual),
        color=curve.color,
        source_curve_id=curve.id,
        x_label=x_label,
        y_label=y_label,
    )


# ──────────────────────────────────────────────────────────
# 辅助
# ──────────────────────────────────────────────────────────

_COLORS = [
    "#0078D4", "#D13438", "#107C10", "#CA5010",
    "#8764B8", "#038387", "#C19C00", "#881798",
]


def _cols_to_series(
    cols: List[List[float]],
    headers: List[str],
    stem: str,
) -> List[DataSeries]:
    """列数据 → DataSeries 列表。

    规则：
    - 1 列 → 1 条系列，x = [0,1,2,...]
    - 2 列 → 1 条系列，x=col0, y=col1
    - N 列（N>2）→ N-1 条系列，共享 col0 为 x
    """
    n = len(cols)
    if n == 0:
        return []
    if n == 1:
        return [DataSeries(
            name=stem,
            x=list(range(len(cols[0]))),
            y=cols[0],
            x_label="index",
            y_label=headers[0],
        )]
    if n == 2:
        return [DataSeries(
            name=stem,
            x=cols[0], y=cols[1],
            x_label=headers[0], y_label=headers[1],
        )]
    # N > 2: x = col0, 余下列各一条
    result = []
    for i in range(1, n):
        color = _COLORS[(i - 1) % len(_COLORS)]
        result.append(DataSeries(
            name=f"{stem}_{headers[i]}",
            x=cols[0], y=cols[i],
            x_label=headers[0], y_label=headers[i],
            color=color,
        ))
    return result


def _detect_sep(line: str) -> Optional[str]:
    candidates = [(",", line.count(",")), ("\t", line.count("\t")), (";", line.count(";"))]
    candidates.sort(key=lambda x: x[1], reverse=True)
    sep, count = candidates[0]
    if count == 0:
        return None  # 空白分隔
    return sep


def _is_header(row: List[str]) -> bool:
    non_numeric = sum(1 for c in row if not _is_float(c))
    return non_numeric > 0


def _is_float(s: str) -> bool:
    try:
        float(s)
        return True
    except ValueError:
        return False
