"""
数据导出模块 - 支持 CSV、Excel、JSON、TXT、剪贴板
"""
from __future__ import annotations

import csv
import json
import math
import os
from html import escape
from pathlib import Path
from typing import List, Optional, TYPE_CHECKING

if TYPE_CHECKING:
    from models.schemas import Curve, DataSeries


class Exporter:
    """曲线数据导出器"""

    _SUPPORTED_SERIES_FORMATS = {"csv", "txt", "dat", "xls", "xlsx"}

    @staticmethod
    def _get_rows(curve: "Curve") -> List[List]:
        """获取数据行 [[x_actual, y_actual], ...]"""
        xs = curve.x_actual if curve.x_actual else curve.x_data
        ys = curve.y_actual if curve.y_actual else curve.y_data
        return list(zip(xs, ys))

    @staticmethod
    def _get_header(curve: "Curve") -> List[str]:
        """根据坐标类型生成表头"""
        if curve.calibration and getattr(curve.calibration, "coord_type", "linear") == "polar":
            return ["θ (角度)", "r (极径)"]
        return ["X", "Y"]

    @staticmethod
    def _get_data_series_rows(series: "DataSeries") -> List[List[float]]:
        rows: List[List[float]] = []
        for x_value, y_value in zip(list(series.x or []), list(series.y or [])):
            try:
                x_num = float(x_value)
                y_num = float(y_value)
            except (TypeError, ValueError):
                continue
            if not (math.isfinite(x_num) and math.isfinite(y_num)):
                continue
            rows.append([x_num, y_num])
        return rows

    @staticmethod
    def _get_data_series_header(series: "DataSeries") -> List[str]:
        x_label = str(getattr(series, "x_label", "") or "X")
        y_label = str(getattr(series, "y_label", "") or "Y")
        return [x_label, y_label]

    @staticmethod
    def _series_payload_from_curve(curve: "Curve") -> dict:
        return {
            "name": curve.name or "curve",
            "header": Exporter._get_header(curve),
            "rows": Exporter._get_rows(curve),
        }

    @staticmethod
    def _series_payload_from_data_series(series: "DataSeries") -> dict:
        return {
            "name": str(series.name or "series"),
            "header": Exporter._get_data_series_header(series),
            "rows": Exporter._get_data_series_rows(series),
        }

    @staticmethod
    def _normalize_data_series_payloads(series_list: List["DataSeries"]) -> List[dict]:
        payloads = [Exporter._series_payload_from_data_series(series) for series in list(series_list or [])]
        payloads = [payload for payload in payloads if payload["rows"]]
        if not payloads:
            raise ValueError("没有可导出的曲线数据")
        return payloads

    @staticmethod
    def _payloads_share_same_x(payloads: List[dict]) -> bool:
        if len(payloads) < 2:
            return False
        base_rows = list(payloads[0].get("rows") or [])
        for payload in payloads[1:]:
            rows = list(payload.get("rows") or [])
            if len(rows) != len(base_rows):
                return False
            for base_row, current_row in zip(base_rows, rows):
                if not math.isclose(float(base_row[0]), float(current_row[0]), rel_tol=0.0, abs_tol=1e-9):
                    return False
        return True

    @staticmethod
    def can_merge_data_series(series_list: List["DataSeries"]) -> bool:
        payloads = Exporter._normalize_data_series_payloads(series_list)
        return Exporter._payloads_share_same_x(payloads)

    @staticmethod
    def _merged_header(payloads: List[dict]) -> List[str]:
        header = [str(payloads[0]["header"][0] or "X")]
        for index, payload in enumerate(payloads, start=1):
            series_name = str(payload.get("name") or "").strip()
            fallback_name = str(payload["header"][1] or f"Y{index}")
            header.append(series_name or fallback_name)
        return header

    @staticmethod
    def _merged_rows(payloads: List[dict]) -> List[List[float]]:
        if not Exporter._payloads_share_same_x(payloads):
            raise ValueError("仅当所有曲线 X 坐标完全对齐时才能合并导出")
        merged_rows: List[List[float]] = []
        row_count = len(payloads[0]["rows"])
        for row_index in range(row_count):
            merged_row = [float(payloads[0]["rows"][row_index][0])]
            for payload in payloads:
                merged_row.append(float(payload["rows"][row_index][1]))
            merged_rows.append(merged_row)
        return merged_rows

    @staticmethod
    def _grouped_delimited_text(
        payloads: List[dict],
        *,
        delimiter: str,
        timestamp: Optional[str] = None,
        merged: bool = False,
    ) -> str:
        lines: List[str] = []
        if timestamp:
            lines.append(f"# exported: {timestamp}")
        if merged:
            header = Exporter._merged_header(payloads)
            lines.append(delimiter.join(header))
            for row in Exporter._merged_rows(payloads):
                lines.append(delimiter.join(str(value) for value in row))
            return "\n".join(lines)
        for index, payload in enumerate(payloads):
            lines.append(f"# {payload['name']}")
            lines.append(delimiter.join(str(value) for value in payload["header"]))
            for row in payload["rows"]:
                lines.append(delimiter.join(str(value) for value in row))
            if index != len(payloads) - 1:
                lines.append("")
        return "\n".join(lines)

    @staticmethod
    def _write_delimited_file(
        payloads: List[dict],
        file_path: str,
        *,
        delimiter: str,
        timestamp: Optional[str] = None,
        merged: bool = False,
    ) -> None:
        if merged:
            rows: List[List[object]] = []
            if timestamp:
                rows.append([f"# exported: {timestamp}"])
            rows.append(Exporter._merged_header(payloads))
            rows.extend(Exporter._merged_rows(payloads))
        else:
            rows = []
            if timestamp:
                rows.append([f"# exported: {timestamp}"])
            for payload in payloads:
                rows.append([f"# {payload['name']}"])
                rows.append(list(payload["header"]))
                rows.extend([list(row) for row in payload["rows"]])
                rows.append([])
            while rows and not rows[-1]:
                rows.pop()
        with open(file_path, "w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.writer(handle, delimiter=delimiter)
            writer.writerows(rows)

    @staticmethod
    def _workbook_sheets(payloads: List[dict], *, timestamp: Optional[str] = None, merged: bool = False) -> List[tuple[str, List[List[object]]]]:
        if merged:
            rows: List[List[object]] = []
            if timestamp:
                rows.append([f"# exported: {timestamp}"])
            rows.append(Exporter._merged_header(payloads))
            rows.extend(Exporter._merged_rows(payloads))
            return [("merged_export", rows)]

        sheets: List[tuple[str, List[List[object]]]] = []
        for index, payload in enumerate(payloads, start=1):
            rows: List[List[object]] = []
            if timestamp:
                rows.append([f"# exported: {timestamp}"])
            rows.append(list(payload["header"]))
            rows.extend([list(row) for row in payload["rows"]])
            sheet_name = str(payload.get("name") or f"series_{index}")
            sheets.append((sheet_name, rows))
        return sheets

    @staticmethod
    def _sanitize_sheet_name(name: str, used_names: set[str], fallback: str) -> str:
        cleaned = str(name or fallback).strip() or fallback
        for char in ("\\", "/", "?", "*", "[", "]", ":"):
            cleaned = cleaned.replace(char, "_")
        cleaned = cleaned[:31] or fallback
        candidate = cleaned
        suffix = 2
        while candidate in used_names:
            base = cleaned[: max(1, 31 - len(str(suffix)) - 1)]
            candidate = f"{base}_{suffix}"
            suffix += 1
        used_names.add(candidate)
        return candidate

    @staticmethod
    def _write_xlsx_file(payloads: List[dict], file_path: str, *, timestamp: Optional[str] = None, merged: bool = False) -> None:
        import openpyxl

        workbook = openpyxl.Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        used_names: set[str] = set()
        for index, (sheet_name, rows) in enumerate(Exporter._workbook_sheets(payloads, timestamp=timestamp, merged=merged), start=1):
            worksheet = workbook.create_sheet(title=Exporter._sanitize_sheet_name(sheet_name, used_names, f"sheet_{index}"))
            for row_index, row in enumerate(rows, start=1):
                for column_index, value in enumerate(row, start=1):
                    worksheet.cell(row=row_index, column=column_index, value=value)
        if not workbook.worksheets:
            workbook.create_sheet("empty")
        workbook.save(file_path)

    @staticmethod
    def _spreadsheet_xml_cell(value: object) -> str:
        if isinstance(value, bool):
            data_type = "String"
            data_value = escape(str(value))
        elif isinstance(value, (int, float)) and math.isfinite(float(value)):
            data_type = "Number"
            data_value = format(float(value), ".12g")
        else:
            data_type = "String"
            data_value = escape(str(value))
        return f'<Cell><Data ss:Type="{data_type}">{data_value}</Data></Cell>'

    @staticmethod
    def _write_xls_xml_file(payloads: List[dict], file_path: str, *, timestamp: Optional[str] = None, merged: bool = False) -> None:
        used_names: set[str] = set()
        lines = [
            '<?xml version="1.0" encoding="UTF-8"?>',
            '<?mso-application progid="Excel.Sheet"?>',
            '<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet"',
            ' xmlns:o="urn:schemas-microsoft-com:office:office"',
            ' xmlns:x="urn:schemas-microsoft-com:office:excel"',
            ' xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet"',
            ' xmlns:html="http://www.w3.org/TR/REC-html40">',
        ]
        for index, (sheet_name, rows) in enumerate(Exporter._workbook_sheets(payloads, timestamp=timestamp, merged=merged), start=1):
            safe_name = Exporter._sanitize_sheet_name(sheet_name, used_names, f"sheet_{index}")
            lines.append(f'  <Worksheet ss:Name="{escape(safe_name)}">')
            lines.append('    <Table>')
            for row in rows:
                lines.append('      <Row>')
                for value in row:
                    lines.append(f"        {Exporter._spreadsheet_xml_cell(value)}")
                lines.append('      </Row>')
            lines.append('    </Table>')
            lines.append('  </Worksheet>')
        lines.append('</Workbook>')
        Path(file_path).write_text("\n".join(lines), encoding="utf-8")

    @staticmethod
    def export_series_file(
        series_list: List["DataSeries"],
        file_path: str,
        *,
        fmt: Optional[str] = None,
        timestamp: Optional[str] = None,
        merged: bool = False,
    ) -> None:
        payloads = Exporter._normalize_data_series_payloads(series_list)
        file_format = str(fmt or Path(file_path).suffix.lstrip(".") or "csv").strip().lower()
        if file_format not in Exporter._SUPPORTED_SERIES_FORMATS:
            raise ValueError(f"不支持的导出格式: {file_format}")
        if merged and not Exporter._payloads_share_same_x(payloads):
            raise ValueError("仅当所有曲线 X 坐标完全对齐时才能合并导出")
        if file_format == "csv":
            Exporter._write_delimited_file(payloads, file_path, delimiter=",", timestamp=timestamp, merged=merged)
            return
        if file_format in {"txt", "dat"}:
            Exporter._write_delimited_file(payloads, file_path, delimiter="\t", timestamp=timestamp, merged=merged)
            return
        if file_format == "xlsx":
            Exporter._write_xlsx_file(payloads, file_path, timestamp=timestamp, merged=merged)
            return
        Exporter._write_xls_xml_file(payloads, file_path, timestamp=timestamp, merged=merged)

    @staticmethod
    def get_series_clipboard_text(
        series_list: List["DataSeries"],
        *,
        timestamp: Optional[str] = None,
        merged: bool = False,
    ) -> str:
        payloads = Exporter._normalize_data_series_payloads(series_list)
        if merged and not Exporter._payloads_share_same_x(payloads):
            raise ValueError("仅当所有曲线 X 坐标完全对齐时才能合并导出")
        return Exporter._grouped_delimited_text(payloads, delimiter="\t", timestamp=timestamp, merged=merged)

    @staticmethod
    def export_series_to_clipboard(
        series_list: List["DataSeries"],
        *,
        timestamp: Optional[str] = None,
        merged: bool = False,
    ) -> None:
        from PySide6.QtWidgets import QApplication

        clipboard = QApplication.clipboard()
        clipboard.setText(Exporter.get_series_clipboard_text(series_list, timestamp=timestamp, merged=merged))

    # ==================== CSV ====================

    @staticmethod
    def export_csv(curve: "Curve", file_path: str, timestamp: Optional[str] = None) -> None:
        """导出为 CSV"""
        rows = Exporter._get_rows(curve)
        header = Exporter._get_header(curve)
        with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            if timestamp:
                writer.writerow([f"# exported: {timestamp}"])
            writer.writerow(["# " + curve.name])
            writer.writerow(header)
            for x, y in rows:
                writer.writerow([x, y])

    @staticmethod
    def export_csv_all(curves: List["Curve"], file_path: str, timestamp: Optional[str] = None) -> None:
        """导出所有曲线到同一 CSV（分组）"""
        with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            if timestamp:
                writer.writerow([f"# exported: {timestamp}"])
            for curve in curves:
                rows = Exporter._get_rows(curve)
                header = Exporter._get_header(curve)
                writer.writerow(["# " + curve.name])
                writer.writerow(header)
                for x, y in rows:
                    writer.writerow([x, y])
                writer.writerow([])

    # ==================== Excel ====================

    @staticmethod
    def export_excel(curve: "Curve", file_path: str, timestamp: Optional[str] = None) -> None:
        """导出为 Excel（单曲线单表）"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = curve.name[:31]  # Excel 表名限制 31 字符

        header = Exporter._get_header(curve)
        rows = Exporter._get_rows(curve)

        start_row = 1
        if timestamp:
            ws.cell(row=1, column=1, value=f"# exported: {timestamp}")
            start_row = 2

        # 表头样式
        header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col, h in enumerate(header, 1):
            cell = ws.cell(row=start_row, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = 16

        # 数据行
        for row_idx, (x, y) in enumerate(rows, start_row + 1):
            ws.cell(row=row_idx, column=1, value=round(x, 8))
            ws.cell(row=row_idx, column=2, value=round(y, 8))

        wb.save(file_path)

    @staticmethod
    def export_excel_all(curves: List["Curve"], file_path: str, timestamp: Optional[str] = None) -> None:
        """导出所有曲线到同一 Excel（每条曲线一个表）"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        # 删除默认 sheet
        default_ws = wb.active
        wb.remove(default_ws)

        for curve in curves:
            ws = wb.create_sheet(title=curve.name[:31])
            header = Exporter._get_header(curve)
            rows = Exporter._get_rows(curve)

            start_row = 1
            if timestamp:
                ws.cell(row=1, column=1, value=f"# exported: {timestamp}")
                start_row = 2

            header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for col, h in enumerate(header, 1):
                cell = ws.cell(row=start_row, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions[cell.column_letter].width = 16

            for row_idx, (x, y) in enumerate(rows, start_row + 1):
                ws.cell(row=row_idx, column=1, value=round(x, 8))
                ws.cell(row=row_idx, column=2, value=round(y, 8))

        if not wb.worksheets:
            wb.create_sheet("空")
        wb.save(file_path)

    # ==================== JSON ====================

    @staticmethod
    def export_json(curve: "Curve", file_path: str, timestamp: Optional[str] = None) -> None:
        """导出为 JSON"""
        rows = Exporter._get_rows(curve)
        header = Exporter._get_header(curve)
        calib = None
        if curve.calibration:
            calib = curve.calibration.model_dump() if hasattr(curve.calibration, "model_dump") else str(curve.calibration)
        data = {
            "name": curve.name,
            "calibration": calib,
            "columns": header,
            "points": [{"x": x, "y": y} for x, y in rows],
        }
        if timestamp:
            data["exported_at"] = timestamp
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

    @staticmethod
    def export_json_all(curves: List["Curve"], file_path: str, timestamp: Optional[str] = None) -> None:
        """导出所有曲线到同一 JSON"""
        all_data = []
        for curve in curves:
            rows = Exporter._get_rows(curve)
            header = Exporter._get_header(curve)
            calib = None
            if curve.calibration:
                calib = curve.calibration.model_dump() if hasattr(curve.calibration, "model_dump") else str(curve.calibration)
            all_data.append({
                "name": curve.name,
                "calibration": calib,
                "columns": header,
                "points": [{"x": x, "y": y} for x, y in rows],
            })
        result = {"curves": all_data}
        if timestamp:
            result["exported_at"] = timestamp
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(result, f, indent=2, ensure_ascii=False)

    # ==================== TXT ====================

    @staticmethod
    def export_txt(curve: "Curve", file_path: str, timestamp: Optional[str] = None) -> None:
        """导出为纯文本（空格分隔）"""
        rows = Exporter._get_rows(curve)
        header = Exporter._get_header(curve)
        with open(file_path, "w", encoding="utf-8") as f:
            if timestamp:
                f.write(f"# exported: {timestamp}\n")
            f.write(f"# {curve.name}\n")
            f.write(f"{header[0]}\t{header[1]}\n")
            for x, y in rows:
                f.write(f"{x}\t{y}\n")

    @staticmethod
    def export_txt_all(curves: List["Curve"], file_path: str, timestamp: Optional[str] = None) -> None:
        """导出所有曲线到同一 TXT"""
        with open(file_path, "w", encoding="utf-8") as f:
            if timestamp:
                f.write(f"# exported: {timestamp}\n")
            for curve in curves:
                rows = Exporter._get_rows(curve)
                header = Exporter._get_header(curve)
                f.write(f"# {curve.name}\n")
                f.write(f"{header[0]}\t{header[1]}\n")
                for x, y in rows:
                    f.write(f"{x}\t{y}\n")
                f.write("\n")

    # ==================== 剪贴板 ====================

    @staticmethod
    def get_clipboard_text(curve: "Curve", timestamp: Optional[str] = None) -> str:
        """生成剪贴板文本（制表符分隔，包含表头）"""
        rows = Exporter._get_rows(curve)
        header = Exporter._get_header(curve)
        lines = []
        if timestamp:
            lines.append(f"# exported: {timestamp}")
        lines.append("\t".join(header))
        for x, y in rows:
            lines.append(f"{x}\t{y}")
        return "\n".join(lines)

    @staticmethod
    def export_to_clipboard(curve: "Curve", timestamp: Optional[str] = None) -> None:
        """将曲线数据复制到系统剪贴板"""
        from PySide6.QtWidgets import QApplication
        text = Exporter.get_clipboard_text(curve, timestamp=timestamp)
        clipboard = QApplication.clipboard()
        clipboard.setText(text)
